"""Maya 2.0 - Excel Tool (read/write .xlsx files via openpyxl, workspace-scoped)"""
from .safe_path import resolve_safe_path


class ExcelTool:
    def read(self, filename: str, sheet: str = None, limit: int = 100, **kwargs) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "Error: openpyxl is not installed on the server"
        try:
            path = resolve_safe_path(filename)
            if not path.exists():
                return f"Error: file not found: {filename}"
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                ws = wb[sheet] if sheet else wb.active
            except KeyError:
                return f"Error: sheet '{sheet}' not found. Available: {', '.join(wb.sheetnames)}"
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= limit:
                    break
                rows.append(row)
            wb.close()
            if not rows:
                return "(empty sheet)"
            return "\n".join(", ".join("" if c is None else str(c) for c in r) for r in rows)
        except ValueError as e:
            return f"Error: {e}"
        except Exception as e:
            return f"Error: could not read Excel file ({e})"

    def write(self, filename: str, rows: list, sheet_name: str = "Sheet1", **kwargs) -> str:
        try:
            from openpyxl import Workbook
        except ImportError:
            return "Error: openpyxl is not installed on the server"
        if not rows:
            return "Error: no rows provided"
        try:
            path = resolve_safe_path(filename)
            path.parent.mkdir(parents=True, exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            for row in rows:
                ws.append(list(row) if isinstance(row, (list, tuple)) else [row])
            wb.save(path)
            return f"Wrote {len(rows)} row(s) to {filename}"
        except ValueError as e:
            return f"Error: {e}"
        except Exception as e:
            return f"Error: could not write Excel file ({e})"
